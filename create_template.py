import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_template_file():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Claims Ingestion Template"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = [
        "Claim Number", 
        "Policy Information", 
        "Accident Date", 
        "Accident Time", 
        "Accident Location", 
        "District State", 
        "Police Station", 
        "Vehicle Registration Numbers", 
        "Vehicle Types", 
        "Involved Parties", 
        "Injury or Death", 
        "Claim Narrative", 
        "Supporting Information"
    ]
    
    # Sample Row
    sample_row = [
        "TP-RCU-UP-00517/2025",
        "POL-998877-2025",
        "2025-05-12",
        "14:30:00",
        "near Kosi Kalan flyover, NH-2",
        "Mathura, Uttar Pradesh",
        "Kosi Kalan PS",
        "UP-85-AT-9988, HR-26-Z-1122",
        "Motorcycle, Truck",
        "Ramesh Kumar, Suresh Singh",
        "Fatal head injuries for rider Ramesh Kumar",
        "The motorcycle UP-85-AT-9988 was hit from behind by the speeding truck HR-26-Z-1122 near Kosi Kalan NH-2 flyover. The rider Ramesh Kumar suffered fatal head injuries.",
        "Previous claim registered for vehicle UP-85-AT-9988 in 2023 for front bumper damage."
    ]
    
    # Styles
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    normal_font = Font(name=font_family, size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # Write Headers
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Write Sample Row
    for col_idx, val in enumerate(sample_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = normal_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border
        
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 60
    
    # Set widths
    for col_idx in range(1, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22
    
    # Save file
    template_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backend", "templates")
    os.makedirs(template_dir, exist_ok=True)
    template_path = os.path.join(template_dir, "claims_upload_template.xlsx")
    wb.save(template_path)
    print(f"Template created successfully at: {template_path}")

if __name__ == "__main__":
    create_template_file()
