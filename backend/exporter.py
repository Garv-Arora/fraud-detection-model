import io
import json
from datetime import datetime
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_evidence_excel(case_data: Dict[str, Any]) -> io.BytesIO:
    """
    Generates a beautifully styled multi-tab Excel spreadsheet for investigator handoff.
    Tabs: Executive Summary, Evidence Links, Image Verification, Audit Log
    """
    wb = openpyxl.Workbook()
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E79")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=12, bold=True, color="1F4E79")
    bold_font = Font(name=font_family, size=11, bold=True)
    normal_font = Font(name=font_family, size=11)
    italic_font = Font(name=font_family, size=9, italic=True, color="595959")
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    accent_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    risk_level = case_data.get("risk_level", "LOW RISK")
    if "HIGH" in risk_level.upper():
        status_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        status_font = Font(name=font_family, size=11, bold=True, color="C00000")
    elif "MEDIUM" in risk_level.upper() or "REVIEW" in risk_level.upper():
        status_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        status_font = Font(name=font_family, size=11, bold=True, color="7F6000")
    else:
        status_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        status_font = Font(name=font_family, size=11, bold=True, color="375623")
        
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # ==========================================
    # SHEET 1: Executive Summary & AI Summary
    # ==========================================
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.cell(row=2, column=2, value="UNIVERSAL SOMPO — AI CLAIM EVIDENCE FINDER REPORT").font = title_font
    
    # Case Details Box
    ws1.cell(row=4, column=2, value="Case Summary").font = section_font
    ws1.cell(row=5, column=2, value="Claim ID").font = bold_font
    ws1.cell(row=5, column=3, value=case_data.get("claim_id")).font = normal_font
    
    ws1.cell(row=6, column=2, value="Policy Information").font = bold_font
    ws1.cell(row=6, column=3, value=case_data.get("policy_information") or "N/A").font = normal_font
    
    ws1.cell(row=7, column=2, value="Risk Level").font = risk_cell_font = status_font
    risk_cell = ws1.cell(row=7, column=3, value=risk_level)
    risk_cell.font = risk_cell_font
    risk_cell.fill = status_fill
    
    ws1.cell(row=8, column=2, value="Overall Score").font = bold_font
    ws1.cell(row=8, column=3, value=case_data.get("overall_score", 0.0)).font = normal_font
    
    ws1.cell(row=9, column=2, value="Quest Pushback Status").font = bold_font
    push_cell = ws1.cell(row=9, column=3, value=case_data.get("pushback_status") or "Not Pushed")
    push_cell.font = bold_font if "Push" in push_cell.value else normal_font
    if "Push" in push_cell.value:
        push_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    ws1.cell(row=10, column=2, value="Processed Date").font = bold_font
    processed_date = case_data.get("created_at")
    if isinstance(processed_date, datetime):
        processed_date = processed_date.strftime("%Y-%m-%d %H:%M:%S")
    ws1.cell(row=10, column=3, value=processed_date).font = normal_font
    
    for r in range(5, 11):
        ws1.cell(row=r, column=2).border = thin_border
        ws1.cell(row=r, column=3).border = thin_border

    # Ingested Facts Table
    ws1.cell(row=12, column=2, value="Confirmed Claim Facts").font = section_font
    facts_fields = [
        ("Accident Date/Time", case_data.get("accident_date_time")),
        ("Loss Location", case_data.get("loss_location")),
        ("Vehicle Numbers", case_data.get("vehicle_numbers")),
        ("Vehicle Types", case_data.get("vehicle_types")),
        ("Parties Involved", case_data.get("parties_involved")),
        ("Injury/Death Summary", case_data.get("injury_or_death")),
        ("Police Station", case_data.get("police_station")),
        ("District & State", case_data.get("district_state")),
        ("FIR Cause Narrative", case_data.get("FIR_cause_narrative")),
        ("Supporting Information", case_data.get("supporting_information"))
    ]
    
    curr_row = 13
    for label, val in facts_fields:
        ws1.cell(row=curr_row, column=2, value=label).font = bold_font
        ws1.cell(row=curr_row, column=2).fill = accent_fill
        ws1.cell(row=curr_row, column=2).border = thin_border
        
        val_cell = ws1.cell(row=curr_row, column=3, value=str(val) if val is not None else "")
        val_cell.font = normal_font
        val_cell.alignment = Alignment(wrap_text=True)
        val_cell.border = thin_border
        curr_row += 1

    # AI Summary Block
    ws1.cell(row=curr_row+1, column=2, value="AI Evidence Summary").font = section_font
    summary_row = curr_row + 2
    ai_sum_cell = ws1.cell(row=summary_row, column=2, value=case_data.get("ai_summary") or "AI Summary pending calculation.")
    ai_sum_cell.font = normal_font
    ai_sum_cell.alignment = Alignment(wrap_text=True, vertical="top")
    # Merge cells to display summary nicely
    ws1.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row+10, end_column=3)
    
    # Border for summary block
    for r in range(summary_row, summary_row + 11):
        for c in range(2, 4):
            ws1.cell(row=r, column=c).border = thin_border
            
    curr_row = summary_row + 12

    # Mismatch Flag Summary
    ws1.cell(row=curr_row, column=2, value="Risk Mismatch Flags").font = section_font
    mismatch_row = curr_row + 1
    
    mismatches = [
        ("CAUSE Mismatch", case_data.get("mismatch_cause")),
        ("LOCATION Mismatch", case_data.get("mismatch_location")),
        ("TIME Mismatch", case_data.get("mismatch_time")),
        ("VEHICLE Mismatch", case_data.get("mismatch_vehicle")),
        ("ENTITY Mismatch", case_data.get("mismatch_entity"))
    ]
    
    ws1.cell(row=mismatch_row, column=2, value="Mismatch Area").font = header_font
    ws1.cell(row=mismatch_row, column=2).fill = header_fill
    ws1.cell(row=mismatch_row, column=3, value="Findings").font = header_font
    ws1.cell(row=mismatch_row, column=3).fill = header_fill
    
    top_mismatches = case_data.get("top_mismatches", "") or ""
    mismatch_list = [m.strip().lower() for m in top_mismatches.split(",") if m.strip()]
    
    mismatch_row += 1
    for area, detail in mismatches:
        ws1.cell(row=mismatch_row, column=2, value=area).font = bold_font
        ws1.cell(row=mismatch_row, column=2).border = thin_border
        
        detail_val = detail if detail else "No mismatch identified."
        detail_cell = ws1.cell(row=mismatch_row, column=3, value=detail_val)
        detail_cell.font = normal_font
        detail_cell.alignment = Alignment(wrap_text=True)
        detail_cell.border = thin_border
        
        area_key = area.split(" ")[0].lower()
        if area_key in mismatch_list:
            ws1.cell(row=mismatch_row, column=2).fill = status_fill
            ws1.cell(row=mismatch_row, column=2).font = Font(name=font_family, size=11, bold=True, color="C00000")
            
        mismatch_row += 1

    ws1.column_dimensions['A'].width = 3
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 85
    
    # ==========================================
    # SHEET 2: Evidence Links (News, Quest, Web, FB, IG)
    # ==========================================
    ws2 = wb.create_sheet(title="Evidence Links")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.cell(row=2, column=2, value="PUBLIC EVIDENCE FINDINGS").font = title_font
    ws2.cell(row=3, column=2, value="Top public articles, Quest histories, Facebook posts, and Instagram media.").font = italic_font
    
    headers2 = ["Rank", "Source Tag", "Title", "URL", "Relevance Score", "Publish Date", "Match Rationale"]
    for col_idx, h in enumerate(headers2, start=2):
        cell = ws2.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    evidences = sorted(case_data.get("evidences", []), key=lambda x: x.get("score", 0), reverse=True)
    
    for idx, ev in enumerate(evidences):
        row_idx = 6 + idx
        ws2.cell(row=row_idx, column=2, value=idx+1).font = normal_font
        
        src_cell = ws2.cell(row=row_idx, column=3, value=ev.get("source"))
        src_cell.alignment = Alignment(horizontal="center")
        src_cell.font = bold_font
        
        # Color coding for V1.0 sources
        src_tag = ev.get("source").upper()
        if "NEWS" in src_tag:
            src_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Blue
        elif "QUEST" in src_tag:
            src_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Yellow
        elif "FACEBOOK" in src_tag:
            src_cell.fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid") # Darker Blue
            src_cell.font = Font(name=font_family, size=11, bold=True, color="1F4E79")
        elif "INSTAGRAM" in src_tag:
            src_cell.fill = PatternFill(start_color="F2CEE6", end_color="F2CEE6", fill_type="solid") # Purple/Pink
            src_cell.font = Font(name=font_family, size=11, bold=True, color="A61C00")
        else:
            src_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Green
            
        ws2.cell(row=row_idx, column=4, value=ev.get("title")).font = normal_font
        
        url_cell = ws2.cell(row=row_idx, column=5, value=ev.get("url"))
        url_cell.font = Font(name=font_family, size=11, color="0563C1", underline="single")
        
        score_cell = ws2.cell(row=row_idx, column=6, value=ev.get("score"))
        score_cell.alignment = Alignment(horizontal="center")
        score_cell.font = bold_font
        
        score_val = ev.get("score", 0.0)
        if score_val >= 0.8:
            score_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        elif score_val >= 0.6:
            score_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            
        ws2.cell(row=row_idx, column=7, value=ev.get("publish_date") or "N/A").font = normal_font
        
        rat_cell = ws2.cell(row=row_idx, column=8, value=ev.get("why_relevant"))
        rat_cell.font = normal_font
        rat_cell.alignment = Alignment(wrap_text=True)
        
        for col_idx in range(2, 9):
            ws2.cell(row=row_idx, column=col_idx).border = thin_border
            
    # Widths
    ws2.column_dimensions['A'].width = 3
    ws2.column_dimensions['B'].width = 6   # Rank
    ws2.column_dimensions['C'].width = 12  # Source Tag
    ws2.column_dimensions['D'].width = 30  # Title
    ws2.column_dimensions['E'].width = 25  # URL
    ws2.column_dimensions['F'].width = 12  # Score
    ws2.column_dimensions['G'].width = 12  # Date
    ws2.column_dimensions['H'].width = 50  # Rationale

    # ==========================================
    # SHEET 3: Google Lens Verification
    # ==========================================
    ws3 = wb.create_sheet(title="Image Verification")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.cell(row=2, column=2, value="GOOGLE LENS PHOTO TRACE").font = title_font
    ws3.cell(row=3, column=2, value="Claim photo reverse lookup results to identify web reuse or stock photo fraud.").font = italic_font
    
    headers3 = ["Photo Name", "Status", "Web Match URL", "Reverse Lookup Details / Match Rationale"]
    for col_idx, h in enumerate(headers3, start=2):
        cell = ws3.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    image_matches = case_data.get("image_matches", [])
    for idx, im in enumerate(image_matches):
        row_idx = 6 + idx
        ws3.cell(row=row_idx, column=2, value=im.get("image_name")).font = bold_font
        
        status = im.get("status")
        status_cell = ws3.cell(row=row_idx, column=3, value=status)
        status_cell.alignment = Alignment(horizontal="center")
        
        if "Stock" in status:
            status_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            status_cell.font = Font(name=font_family, size=11, bold=True, color="C00000")
        elif "Prior" in status:
            status_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            status_cell.font = Font(name=font_family, size=11, bold=True, color="7F6000")
        else:
            status_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            status_cell.font = Font(name=font_family, size=11, bold=True, color="375623")
            
        url_cell = ws3.cell(row=row_idx, column=4, value=im.get("matched_url") or "No Matches")
        if im.get("matched_url"):
            url_cell.font = Font(name=font_family, size=11, color="0563C1", underline="single")
        else:
            url_cell.font = normal_font
            
        why_cell = ws3.cell(row=row_idx, column=5, value=im.get("why_matched"))
        why_cell.font = normal_font
        why_cell.alignment = Alignment(wrap_text=True)
        
        for col_idx in range(2, 6):
            ws3.cell(row=row_idx, column=col_idx).border = thin_border
            
    # Widths
    ws3.column_dimensions['A'].width = 3
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 35
    ws3.column_dimensions['E'].width = 50

    # ==========================================
    # SHEET 4: System Audit Log
    # ==========================================
    ws4 = wb.create_sheet(title="System Audit Log")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4.cell(row=2, column=2, value="INVESTIGATION AUDIT LOG").font = title_font
    ws4.cell(row=3, column=2, value="Traceable actions, query generation expansions, and investigator updates.").font = italic_font
    
    headers4 = ["Timestamp (UTC)", "Action", "Details / Parameters Used"]
    for col_idx, h in enumerate(headers4, start=2):
        cell = ws4.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    audit_logs = case_data.get("audit_logs", [])
    audit_logs = sorted(audit_logs, key=lambda x: x.get("timestamp") or datetime.min)
    
    for idx, al in enumerate(audit_logs):
        row_idx = 6 + idx
        ts = al.get("timestamp")
        if isinstance(ts, datetime):
            ts = ts.strftime("%Y-%m-%d %H:%M:%S")
        ws4.cell(row=row_idx, column=2, value=ts).font = normal_font
        ws4.cell(row=row_idx, column=3, value=al.get("action")).font = bold_font
        
        details_cell = ws4.cell(row=row_idx, column=4, value=al.get("details"))
        details_cell.font = normal_font
        details_cell.alignment = Alignment(wrap_text=True)
        
        for col_idx in range(2, 5):
            ws4.cell(row=row_idx, column=col_idx).border = thin_border
            
    # Widths
    ws4.column_dimensions['A'].width = 3
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 75
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

def generate_evidence_html_report(case_data: Dict[str, Any]) -> str:
    """Generates a premium print-ready HTML page representing the PDF evidence report."""
    claim_id = case_data.get("claim_id")
    risk_level = case_data.get("risk_level", "LOW RISK")
    overall_score = case_data.get("overall_score", 0.0)
    policy_info = case_data.get("policy_information") or "N/A"
    pushback = case_data.get("pushback_status") or "Not Pushed"
    
    risk_class = "risk-low"
    if "HIGH" in risk_level.upper():
        risk_class = "risk-high"
    elif "MEDIUM" in risk_level.upper() or "REVIEW" in risk_level.upper():
        risk_class = "risk-medium"
        
    facts_html = ""
    fields = [
        ("Policy Information", policy_info),
        ("Accident Date/Time", case_data.get("accident_date_time")),
        ("Loss Location", case_data.get("loss_location")),
        ("Vehicle Numbers", case_data.get("vehicle_numbers")),
        ("Vehicle Types", case_data.get("vehicle_types")),
        ("Parties Involved", case_data.get("parties_involved")),
        ("Injury/Death Summary", case_data.get("injury_or_death")),
        ("Police Station", case_data.get("police_station")),
        ("District & State", case_data.get("district_state")),
        ("FIR Cause Narrative", case_data.get("FIR_cause_narrative")),
        ("Supporting Information", case_data.get("supporting_information"))
    ]
    for label, val in fields:
        facts_html += f"""
        <tr>
            <td class="fact-label">{label}</td>
            <td class="fact-value">{val if val is not None else ""}</td>
        </tr>
        """
        
    mismatches = [
        ("Cause", case_data.get("mismatch_cause")),
        ("Location", case_data.get("mismatch_location")),
        ("Time", case_data.get("mismatch_time")),
        ("Vehicle", case_data.get("mismatch_vehicle")),
        ("Entity", case_data.get("mismatch_entity"))
    ]
    mismatch_html = ""
    top_mismatches = case_data.get("top_mismatches", "") or ""
    mismatch_list = [m.strip().lower() for m in top_mismatches.split(",") if m.strip()]
    
    for area, detail in mismatches:
        is_flagged = area.lower() in mismatch_list
        flag_status = '<span class="badge badge-error">FLAGGED MISMATCH</span>' if is_flagged else '<span class="badge badge-success">MATCH OK</span>'
        mismatch_html += f"""
        <div class="mismatch-card {'flagged' if is_flagged else ''}">
            <div class="mismatch-header">
                <strong>{area} Check</strong>
                {flag_status}
            </div>
            <div class="mismatch-body">{detail or "No discrepancies identified between public evidence and claim facts."}</div>
        </div>
        """
        
    evidences = sorted(case_data.get("evidences", []), key=lambda x: x.get("score", 0), reverse=True)
    evidence_rows = ""
    for idx, ev in enumerate(evidences):
        score = ev.get("score", 0.0)
        score_class = "score-high" if score >= 0.8 else ("score-medium" if score >= 0.6 else "score-low")
        evidence_rows += f"""
        <tr>
            <td style="text-align: center;">{idx+1}</td>
            <td><span class="source-tag tag-{ev.get('source').lower()}">{ev.get('source')}</span></td>
            <td><strong>{ev.get('title')}</strong><br><span class="url-text">{ev.get('url')}</span></td>
            <td style="text-align: center;"><span class="score-badge {score_class}">{score:.2f}</span></td>
            <td>{ev.get('publish_date') or 'N/A'}</td>
            <td>{ev.get('why_relevant')}</td>
        </tr>
        """
        
    images_html = ""
    image_matches = case_data.get("image_matches", [])
    for im in image_matches:
        status = im.get("status")
        status_class = "status-original"
        if "Stock" in status:
            status_class = "status-stock"
        elif "Prior" in status:
            status_class = "status-prior"
            
        images_html += f"""
        <div class="image-card">
            <div class="image-title">
                <strong>{im.get('image_name')}</strong>
                <span class="image-status {status_class}">{status}</span>
            </div>
            <div class="image-details">
                {f'<strong>Matched URL:</strong> <a href="{im.get("matched_url")}" target="_blank">{im.get("matched_url")}</a><br>' if im.get("matched_url") else ''}
                <strong>Analysis:</strong> {im.get('why_matched')}
            </div>
        </div>
        """
        
    audit_rows = ""
    audit_logs = sorted(case_data.get("audit_logs", []), key=lambda x: x.get("timestamp") or datetime.min)
    for al in audit_logs:
        ts = al.get("timestamp")
        if isinstance(ts, datetime):
            ts = ts.strftime("%Y-%m-%d %H:%M:%S")
        audit_rows += f"""
        <tr>
            <td style="white-space: nowrap;">{ts}</td>
            <td><strong>{al.get('action')}</strong></td>
            <td>{al.get('details')}</td>
        </tr>
        """

    # HTML formatted AI summary block
    # Convert markdown-like symbols in heuristic summaries to HTML linebreaks
    ai_summary_raw = case_data.get("ai_summary") or "AI Summary pending calculation."
    ai_summary_html = ai_summary_raw.replace("\n", "<br>").replace("### ", "<h4>").replace("###", "</h4>").replace("* **", "<strong>").replace("**: ", "</strong>: ").replace("* ", "• ")

    html = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <title>Claim Evidence Report — {claim_id}</title>
        <style>
            body {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                color: #333;
                line-height: 1.5;
                padding: 30px;
                background-color: #fff;
            }}
            .header-table {{
                width: 100%;
                border-bottom: 3px solid #1F4E79;
                padding-bottom: 10px;
                margin-bottom: 30px;
            }}
            .header-title {{
                font-size: 24px;
                font-weight: bold;
                color: #1F4E79;
                margin: 0;
            }}
            .header-subtitle {{
                font-size: 14px;
                color: #595959;
                margin: 5px 0 0 0;
            }}
            .section-title {{
                font-size: 18px;
                font-weight: bold;
                color: #1F4E79;
                margin-top: 30px;
                margin-bottom: 15px;
                border-bottom: 1px solid #BFBFBF;
                padding-bottom: 5px;
            }}
            
            .risk-badge {{
                display: inline-block;
                padding: 6px 14px;
                border-radius: 4px;
                font-weight: bold;
                text-align: center;
                font-size: 14px;
            }}
            .risk-high {{ background-color: #FCE4D6; color: #C00000; border: 1px solid #C00000; }}
            .risk-medium {{ background-color: #FFF2CC; color: #7F6000; border: 1px solid #7F6000; }}
            .risk-low {{ background-color: #E2EFDA; color: #375623; border: 1px solid #375623; }}
            
            .badge {{
                display: inline-block;
                padding: 2px 8px;
                border-radius: 3px;
                font-size: 10px;
                font-weight: bold;
            }}
            .badge-error {{ background-color: #FCE4D6; color: #C00000; }}
            .badge-success {{ background-color: #E2EFDA; color: #375623; }}
            
            .source-tag {{
                display: inline-block;
                padding: 2px 8px;
                border-radius: 4px;
                font-size: 11px;
                font-weight: bold;
                color: #333;
                text-align: center;
            }}
            .tag-news {{ background-color: #DDEBF7; }}
            .tag-quest {{ background-color: #FFF2CC; }}
            .tag-facebook {{ background-color: #C9DAF8; color: #1F4E79; }}
            .tag-instagram {{ background-color: #F2CEE6; color: #A61C00; }}
            .tag-image {{ background-color: #FCE4D6; }}
            .tag-web {{ background-color: #E2EFDA; }}
            
            .score-badge {{
                display: inline-block;
                padding: 3px 6px;
                border-radius: 3px;
                font-weight: bold;
                font-size: 12px;
            }}
            .score-high {{ background-color: #E2EFDA; color: #375623; }}
            .score-medium {{ background-color: #FFF2CC; color: #7F6000; }}
            .score-low {{ background-color: #FCE4D6; color: #C00000; }}
            
            .summary-container {{
                display: flex;
                gap: 30px;
                margin-bottom: 30px;
            }}
            .summary-box {{
                flex: 1;
                border: 1px solid #BFBFBF;
                border-radius: 6px;
                padding: 15px;
                background-color: #F9FAFB;
            }}
            .summary-table {{
                width: 100%;
                border-collapse: collapse;
            }}
            .summary-table td {{
                padding: 6px 10px;
                vertical-align: top;
            }}
            
            .ai-summary-block {{
                background-color: #F8F9FA;
                border-left: 4px solid #1F4E79;
                padding: 20px;
                border-radius: 4px;
                margin-bottom: 30px;
                font-size: 13.5px;
            }}
            .ai-summary-block h4 {{
                color: #1F4E79;
                margin-top: 15px;
                margin-bottom: 8px;
            }}
            .ai-summary-block h4:first-child {{
                margin-top: 0;
            }}
            
            .facts-table {{
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 30px;
                border: 1px solid #BFBFBF;
            }}
            .facts-table td {{
                padding: 8px 12px;
                border: 1px solid #BFBFBF;
                vertical-align: top;
            }}
            .fact-label {{
                width: 25%;
                font-weight: bold;
                background-color: #F2F4F7;
            }}
            .fact-value {{
                width: 75%;
            }}
            
            .mismatch-grid {{
                display: grid;
                grid-template-columns: 1fr;
                gap: 15px;
                margin-bottom: 30px;
            }}
            .mismatch-card {{
                border: 1px solid #E4E7EC;
                border-radius: 6px;
                padding: 15px;
                background-color: #FCFCFD;
            }}
            .mismatch-card.flagged {{
                border: 1px solid #FDA29B;
                background-color: #FEF3F2;
            }}
            .mismatch-header {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                margin-bottom: 8px;
            }}
            .mismatch-body {{
                font-size: 13px;
                color: #475467;
            }}
            
            .data-table {{
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 30px;
            }}
            .data-table th {{
                background-color: #1F4E79;
                color: #ffffff;
                padding: 10px 12px;
                font-weight: bold;
                border: 1px solid #1F4E79;
                font-size: 12px;
            }}
            .data-table td {{
                padding: 8px 12px;
                border: 1px solid #EAECF0;
                font-size: 12px;
                vertical-align: top;
            }}
            .url-text {{
                font-size: 10px;
                color: #595959;
                word-break: break-all;
            }}
            
            .image-grid {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 20px;
                margin-bottom: 30px;
            }}
            .image-card {{
                border: 1px solid #EAECF0;
                border-radius: 6px;
                padding: 15px;
                background-color: #F9FAFB;
            }}
            .image-title {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                margin-bottom: 10px;
            }}
            .image-status {{
                font-size: 11px;
                font-weight: bold;
                padding: 2px 8px;
                border-radius: 3px;
            }}
            .status-stock {{ background-color: #FCE4D6; color: #C00000; }}
            .status-prior {{ background-color: #FFF2CC; color: #7F6000; }}
            .status-original {{ background-color: #E2EFDA; color: #375623; }}
            .image-details {{
                font-size: 12px;
                color: #475467;
            }}
            
            @media print {{
                body {{
                    padding: 0;
                    background-color: #fff;
                    font-size: 12px;
                }}
                .mismatch-card.flagged {{
                    background-color: #FFF0F0 !important;
                    -webkit-print-color-adjust: exact;
                    print-color-adjust: exact;
                }}
                .data-table th {{
                    background-color: #1F4E79 !important;
                    color: #fff !important;
                    -webkit-print-color-adjust: exact;
                    print-color-adjust: exact;
                }}
                .source-tag, .score-badge, .risk-badge, .badge {{
                    -webkit-print-color-adjust: exact;
                    print-color-adjust: exact;
                }}
                .no-print {{
                    display: none;
                }}
            }}
            
            .btn-print {{
                background-color: #1F4E79;
                color: white;
                border: none;
                padding: 8px 16px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 4px;
                cursor: pointer;
                float: right;
                margin-top: 10px;
            }}
            .btn-print:hover {{
                background-color: #153553;
            }}
        </style>
    </head>
    <body>
        <button class="btn-print no-print" onclick="window.print()">Print Report / Save PDF</button>
        
        <table class="header-table">
            <tr>
                <td>
                    <div class="header-title">UNIVERSAL SOMPO — FRAUD DISCOVERY REPORT</div>
                    <div class="header-subtitle">AI Claim Evidence Finder Pilot Dashboard (V1.0 Scope)</div>
                </td>
                <td style="text-align: right; vertical-align: middle;">
                    <span class="risk-badge {risk_class}">{risk_level}</span>
                </td>
            </tr>
        </table>
        
        <div class="summary-container">
            <div class="summary-box">
                <h3 style="margin-top:0; color:#1F4E79;">Case Details</h3>
                <table class="summary-table">
                    <tr>
                        <td style="font-weight: bold; width: 45%;">Claim ID:</td>
                        <td>{claim_id}</td>
                    </tr>
                    <tr>
                        <td style="font-weight: bold;">Policy ID:</td>
                        <td>{policy_info}</td>
                    </tr>
                    <tr>
                        <td style="font-weight: bold;">Risk Level:</td>
                        <td><strong>{risk_level}</strong></td>
                    </tr>
                    <tr>
                        <td style="font-weight: bold;">Overall Evidence Score:</td>
                        <td>{overall_score:.2f}</td>
                    </tr>
                </table>
            </div>
            
            <div class="summary-box">
                <h3 style="margin-top:0; color:#1F4E79;">Metadata</h3>
                <table class="summary-table">
                    <tr>
                        <td style="font-weight: bold; width: 45%;">Processed On:</td>
                        <td>{case_data.get('created_at').strftime('%Y-%m-%d %H:%M:%S') if isinstance(case_data.get('created_at'), datetime) else case_data.get('created_at')}</td>
                    </tr>
                    <tr>
                        <td style="font-weight: bold;">Quest Pushback:</td>
                        <td><span style="font-weight: bold; color: { '#27ae60' if 'Push' in pushback else '#333' }">{pushback}</span></td>
                    </tr>
                    <tr>
                        <td style="font-weight: bold;">Evidence Found:</td>
                        <td>{len(evidences)} items</td>
                    </tr>
                </table>
            </div>
        </div>
        
        <div class="section-title">AI Evidence Summary</div>
        <div class="ai-summary-block">
            {ai_summary_html}
        </div>
        
        <div class="section-title">Ingested Case Facts</div>
        <table class="facts-table">
            {facts_html}
        </table>
        
        <div class="section-title">Risk Mismatch Analysis</div>
        <div class="mismatch-grid">
            {mismatch_html}
        </div>
        
        <div class="section-title">Ranked Public Evidence Sources</div>
        <table class="data-table">
            <thead>
                <tr>
                    <th style="width: 5%;">Rank</th>
                    <th style="width: 12%;">Source</th>
                    <th style="width: 33%;">Title / URL</th>
                    <th style="width: 10%;">Score</th>
                    <th style="width: 12%;">Publish Date</th>
                    <th style="width: 28%;">Match Rationale</th>
                </tr>
            </thead>
            <tbody>
                {evidence_rows}
            </tbody>
        </table>
        
        <div class="section-title">Google Lens Photo Verification</div>
        <div class="image-grid">
            {images_html if image_matches else '<div style="grid-column: span 2; color:#595959; font-style:italic;">No claim photos uploaded for visual verification.</div>'}
        </div>
        
        <div class="section-title">System Audit Log</div>
        <table class="data-table" style="margin-bottom:0;">
            <thead>
                <tr>
                    <th style="width: 20%;">Timestamp (UTC)</th>
                    <th style="width: 20%;">Action</th>
                    <th style="width: 60%;">Details</th>
                </tr>
            </thead>
            <tbody>
                {audit_rows}
            </tbody>
        </table>
        
        <div style="margin-top: 40px; border-top: 1px solid #BFBFBF; padding-top: 10px; font-size: 10px; color:#595959; text-align: center;">
             Universal Sompo AI Claim Evidence Finder — Statement of Responsibility: The platform is a decision support tool that does not make claim approval, rejection, or fraud decisions. All final actions are the responsibility of authorized investigators.
        </div>
    </body>
    </html>
    """
    return html
