import os
import io
import unittest
import openpyxl
from fastapi.testclient import TestClient
from backend.main import app
import backend.extractor as extractor

class TestUniversalSompoTeraBotIntegration(unittest.TestCase):
    def setUp(self):
        self.client = TestClient(app)

    def test_01_intimation_parser(self):
        sample_zip = os.path.join("samples", "CL26148443.zip")
        self.assertTrue(os.path.exists(sample_zip), "samples/CL26148443.zip must exist")
        
        facts, conf = extractor.extract_facts_from_zip(sample_zip)
        self.assertEqual(facts.get("claim_id"), "CL26148443")
        self.assertEqual(facts.get("insured_name"), "Manju")
        self.assertEqual(facts.get("vehicle_make"), "MARUTI")
        self.assertIn("RJ-45-CQ-1390", facts.get("vehicle_numbers"))
        self.assertEqual(facts.get("accident_location_city"), "MUMBAI")

    def test_02_load_sample_presets_endpoint(self):
        res = self.client.post("/api/cases/load-sample-presets")
        self.assertEqual(res.status_code, 200)
        data = res.json()
        self.assertTrue(data.get("success"))
        self.assertGreaterEqual(len(data.get("claims", [])), 4)

    def test_03_list_cases(self):
        res = self.client.get("/api/cases")
        self.assertEqual(res.status_code, 200)
        cases = res.json()
        self.assertGreaterEqual(len(cases), 4)
        
        # Check 30-header properties exist on returned case
        c = cases[0]
        self.assertIn("claim_id", c)
        self.assertIn("insured_name", c)
        self.assertIn("vehicle_make", c)
        self.assertIn("news_check", c)
        self.assertIn("social_media_check", c)

    def test_04_export_excel_terabot_tab(self):
        # Fetch one case ID
        res = self.client.get("/api/cases")
        cases = res.json()
        claim_id = cases[0]["claim_id"]
        
        export_res = self.client.get(f"/api/cases/{claim_id}/export-excel")
        self.assertEqual(export_res.status_code, 200)
        
        # Parse Excel stream
        wb = openpyxl.load_workbook(filename=io.BytesIO(export_res.content))
        self.assertIn("Tera Bot 30 Headers", wb.sheetnames)
        
        ws = wb["Tera Bot 30 Headers"]
        # Verify 30 header columns in Row 4
        headers = [ws.cell(row=4, column=c).value for c in range(1, 31)]
        self.assertEqual(len(headers), 30)
        self.assertEqual(headers[0], "Claim No ")
        self.assertEqual(headers[1], "Insured Name")
        self.assertEqual(headers[4], "Vehicle No")

    def test_05_upload_excel_terabot_file(self):
        excel_path = os.path.join("samples", "Headers - Tera Bot.xlsx")
        self.assertTrue(os.path.exists(excel_path), f"{excel_path} must exist")
        
        with open(excel_path, "rb") as f:
            res = self.client.post("/api/cases/upload-excel", files={"file": ("Headers - Tera Bot.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        self.assertEqual(res.status_code, 200)
        data = res.json()
        self.assertTrue(data.get("success"))
        self.assertGreaterEqual(len(data.get("claims", [])), 20)

    def test_06_vehicle_permutations_generator(self):
        import backend.search_engine as search_engine
        perms = search_engine.generate_vehicle_permutations("RJ-09-GC-8889")
        self.assertIn("RJ-09-GC-8889", perms)
        self.assertIn("RJ09GC8889", perms)
        self.assertIn("RJ 09 GC 8889", perms)
        self.assertIn("8889", perms)

    def test_07_vernacular_and_social_query_generation(self):
        import backend.search_engine as search_engine
        facts = {
            "claim_id": "CL26121725",
            "insured_name": "Arun Kumar Pal",
            "driver_name": "Arun Kumar Pal",
            "vehicle_numbers": ["UP-66-K-9912"],
            "loss_location": "Durgaganj, Suriyawan",
            "district_state": "Bhadohi, Uttar Pradesh",
            "accident_date_time": "2026-05-01T21:30:00",
            "FIR_cause_narrative": "Car collided with truck near Durgaganj PS"
        }
        queries = search_engine.generate_search_queries(facts)
        self.assertGreaterEqual(len(queries), 5)
        # Verify vernacular / social media query generation
        query_text = " ".join(queries)
        self.assertTrue("सड़क दुर्घटना" in query_text or "बारात" in query_text or "instagram" in query_text)

    def test_08_repudiated_benchmark_mismatch_evaluations(self):
        import backend.scorer as scorer
        
        # Test Case 1: Driver Implant (Case CL21246240)
        facts1 = {"driver_name": "Naushad", "insured_name": "Dinesh Kumar"}
        evs1 = [{"title": "News Report", "snippet": "Sushil was driving the vehicle without DL", "score": 0.85}]
        flags1, exp1 = scorer.evaluate_mismatch_flags(facts1, evs1)
        self.assertIn("driver_implant", flags1)

        # Test Case 2: Pre-Inception Loss (Case CL24181742)
        facts2 = {"accident_date_time": "2024-07-14", "insured_name": "Chanda Chhabra"}
        evs2 = [{"title": "Instagram Video", "snippet": "Video uploaded on 11.07.2024 shows damage prior to the policy start date", "score": 0.90}]
        flags2, exp2 = scorer.evaluate_mismatch_flags(facts2, evs2)
        self.assertIn("pre_inception", flags2)

        # Test Case 3: Wedding Barat / Hire & Reward (Case CL26121725)
        facts3 = {"insured_name": "Arun Kumar Pal", "loss_location": "Durgaganj"}
        evs3 = [{"title": "Dainik Bhaskar", "snippet": "Car going in Wedding Procession (Barat) with Groom Manjit Pal collided with truck", "score": 0.92}]
        flags3, exp3 = scorer.evaluate_mismatch_flags(facts3, evs3)
        self.assertIn("hire_and_reward", flags3)

    def test_09_custom_search_endpoint(self):
        res = self.client.post("/api/cases/CL26121725/custom-search", json={"queries": ["Durgaganj barat accident", "UP66K9912 accident"]})
        self.assertEqual(res.status_code, 200)
        data = res.json()
        self.assertTrue(data.get("success"))
        self.assertEqual(len(data.get("queries")), 2)

if __name__ == "__main__":
    unittest.main()

